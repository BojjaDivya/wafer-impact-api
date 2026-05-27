"""
Parser for Excel files in the simplified BrightWay2 format used by both
Company A and partner companies.
"""

from openpyxl import load_workbook
from typing import IO

from .models import Activity, Exchange


def parse_excel_database(source) -> dict:
    """Parse an Excel file; return dict of activities keyed by name."""
    wb = load_workbook(source, read_only=True, data_only=True)
    return _parse_bw_sheet(wb["BW database"])


def parse_full_database(source) -> tuple:
    """
    Parse a full Company A database Excel file.
    Returns (activities, materials, electricity).
    """
    wb = load_workbook(source, read_only=True, data_only=True)
    activities = _parse_bw_sheet(wb["BW database"])
    materials = _parse_impact_sheet(wb["Materials"])
    electricity = _parse_impact_sheet(wb["Electricity"])
    return activities, materials, electricity


def _parse_bw_sheet(ws) -> dict:
    """
    Parse the 'BW database' sheet.

    Layout per activity:
      Activity    <name>
      location    <loc>
      unit        <unit>
      reference product  <ref>
      type        process
      (blank)
      Exchanges
      name  amount  unit  location  type  reference product
      <exchange rows ...>
      (blank blank between activities)
    """
    activities: dict = {}
    current_activity = None
    in_exchanges = False

    for row in ws.iter_rows(values_only=True):
        first = str(row[0]).strip() if row[0] is not None else ""
        all_blank = all(cell is None for cell in row)

        if all_blank:
            # Blank rows inside a block are fine; reset ONLY when NOT in
            # an exchange block (consecutive blanks between activities).
            # We actually just keep going — "Activity" row resets naturally.
            if not in_exchanges:
                pass  # keep current_activity; next "Activity" will reset
            continue

        if first == "Activity":
            name = str(row[1]).strip()
            current_activity = Activity(
                name=name, location="", unit="", reference_product=""
            )
            activities[name] = current_activity
            in_exchanges = False

        elif first == "location" and current_activity:
            current_activity.location = str(row[1]).strip() if row[1] else ""

        elif first == "unit" and current_activity:
            current_activity.unit = str(row[1]).strip() if row[1] else ""

        elif first == "reference product" and current_activity:
            current_activity.reference_product = str(row[1]).strip() if row[1] else ""

        elif first == "Exchanges" and current_activity:
            in_exchanges = True  # next row is the column header

        elif in_exchanges and first == "name":
            pass  # skip column header row

        elif in_exchanges and current_activity and first:
            # Exchange data row: name, amount, unit, location, type, ref product
            try:
                exc = Exchange(
                    name=str(row[0]).strip(),
                    amount=float(row[1]) if row[1] is not None else 0.0,
                    unit=str(row[2]).strip() if row[2] else "",
                    location=str(row[3]).strip() if row[3] else "",
                    type=str(row[4]).strip() if row[4] else "technosphere",
                    reference_product=str(row[5]).strip() if row[5] else "",
                )
                current_activity.exchanges.append(exc)
            except (ValueError, TypeError):
                pass

    return activities


def _parse_impact_sheet(ws) -> dict:
    """Parse Materials or Electricity sheet. First row is header."""
    factors = {}
    first_row = True
    for row in ws.iter_rows(values_only=True):
        if first_row:
            first_row = False
            continue
        if row[0] is None:
            continue
        name = str(row[0]).strip()
        try:
            factors[name] = float(row[1])
        except (TypeError, ValueError):
            pass
    return factors
